from __future__ import annotations

import os
from copy import copy
from decimal import Decimal, ROUND_HALF_UP, InvalidOperation
from pathlib import Path
from typing import Any
import json

import openpyxl
import requests
from openpyxl.utils import get_column_letter
import config

API_URL = "https://api-seller.ozon.ru/v5/product/info/prices"



OZON_CLIENT_ID = config.OZON_ID
OZON_API_KEY = config.OZON_TOKEN


def q2(value: Any) -> Decimal:
    """
    Математическое округление до 2 знаков.
    """
    try:
        return Decimal(str(value)).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
    except (InvalidOperation, ValueError, TypeError):
        return Decimal("0.00")


def normalize(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip().lower()


def chunked(seq: list[Any], size: int):
    for i in range(0, len(seq), size):
        yield seq[i:i + size]


def find_header_map(ws) -> dict[str, int]:
    headers = {}
    for col in range(1, ws.max_column + 1):
        headers[normalize(ws.cell(1, col).value)] = col
    return headers


def parse_product_id(value: Any) -> str | None:
    if value is None:
        return None

    s = str(value).strip()
    if not s:
        return None

    try:
        # На случай, если Excel хранит 12345 как 12345.0
        return str(int(float(s)))
    except Exception:
        return None


def parse_offer_id(value: Any) -> str | None:
    if value is None:
        return None

    s = str(value).strip()
    return s or None


def copy_cell_style(src_cell, dst_cell) -> None:
    if src_cell.has_style:
        dst_cell._style = copy(src_cell._style)

    dst_cell.font = copy(src_cell.font)
    dst_cell.fill = copy(src_cell.fill)
    dst_cell.border = copy(src_cell.border)
    dst_cell.alignment = copy(src_cell.alignment)
    dst_cell.protection = copy(src_cell.protection)
    dst_cell.number_format = src_cell.number_format


def add_new_columns(ws, headers: list[str]) -> list[int]:
    """
    Добавляет новые колонки справа и возвращает их индексы.
    """
    base_col = ws.max_column
    start_col = base_col + 1
    added_cols = []

    for i, header in enumerate(headers):
        col = start_col + i
        added_cols.append(col)
        ws.cell(1, col).value = header

        # Заголовки копируем по стилю из последней исходной колонки
        copy_cell_style(ws.cell(1, base_col), ws.cell(1, col))

        # И строки тоже
        for row in range(2, ws.max_row + 1):
            copy_cell_style(ws.cell(row, base_col), ws.cell(row, col))

    return added_cols


def build_output_path(input_file: str | Path) -> Path:
    input_path = Path(input_file)
    return input_path.with_name(f"{input_path.stem}_modified{input_path.suffix}")


def fetch_prices(
    ids: list[str],
    id_field: str,  # "product_id" или "offer_id"
    client_id: str,
    api_key: str,
    verbose: bool = True,
) -> dict[str, Decimal]:
    """
    Возвращает словарь:
      key -> price
    где key это строковый product_id или offer_id.
    """
    result: dict[str, Decimal] = {}

    ids = [str(x).strip() for x in ids if str(x).strip()]
    ids = list(dict.fromkeys(ids))  # убрать дубли, сохранить порядок

    if not ids:
        if verbose:
            print(f"[{id_field}] Пустой список идентификаторов, запрос не нужен.")
        return result

    headers = {
        "Client-Id": client_id,
        "Api-Key": api_key,
        "Content-Type": "application/json",
    }

    if verbose:
        print(f"[{id_field}] Уникальных идентификаторов: {len(ids)}")
        print(f"[{id_field}] Первые 10 идентификаторов: {ids[:10]}")

    with requests.Session() as session:
        for batch_no, batch in enumerate(chunked(ids, 1000), start=1):
            cursor = ""
            page_no = 1

            while True:
                payload = {
                    "cursor": cursor,
                    "filter": {
                        id_field: batch,
                        "visibility": "ALL",
                    },
                    "limit": 1000,
                }

                if verbose:
                    print()
                    print(f"[{id_field}] Батч {batch_no}, страница {page_no}")
                    print(f"[{id_field}] Размер батча: {len(batch)}")
                    print(f"[{id_field}] cursor: {repr(cursor)}")
                    print(f"[{id_field}] payload filter keys: {list(payload['filter'].keys())}")

                try:
                    response = session.post(API_URL, headers=headers, json=payload, timeout=60)
                except Exception as e:
                    print(f"[{id_field}] Ошибка HTTP-запроса: {e}")
                    raise

                if verbose:
                    print(f"[{id_field}] HTTP status: {response.status_code}")

                if not response.ok:
                    print(f"[{id_field}] Текст ошибки Ozon:")
                    print(response.text[:5000])
                    response.raise_for_status()

                try:
                    data = response.json()
                except Exception:
                    print(f"[{id_field}] Не удалось распарсить JSON. Ответ:")
                    print(response.text[:5000])
                    raise

                items = data.get("items", [])
                next_cursor = data.get("cursor") or ""
                total = data.get("total")

                if verbose:
                    print(f"[{id_field}] items в ответе: {len(items)}")
                    print(f"[{id_field}] total: {total}")
                    print(f"[{id_field}] next_cursor пустой?: {not bool(next_cursor)}")

                    if items:
                        example = items[0]
                        print(f"[{id_field}] Пример первого item:")
                        print(json.dumps(example, ensure_ascii=False, indent=2)[:4000])

                for item in items:
                    key_raw = item.get(id_field)
                    price_obj = item.get("price") or {}
                    price_raw = price_obj.get("price")

                    if key_raw in (None, "") or price_raw in (None, ""):
                        continue

                    key = str(key_raw).strip()
                    result[key] = q2(price_raw)

                if not next_cursor:
                    break

                cursor = next_cursor
                page_no += 1

    if verbose:
        print(f"[{id_field}] Итого цен получено: {len(result)}")

    return result


def enrich_ozon_prices_xlsx(
    input_file: str | Path,
    byn_rate_per_100_rub: float | str | Decimal,
    sheet_name: str | None = None,
    client_id: str | None = None,
    api_key: str | None = None,
    verbose: bool = True,
) -> Path:
    """
    Принимает:
      - input_file: путь к xlsx
      - byn_rate_per_100_rub: курс BYN к 100 RUB, например 3.7677

    Создаёт новый файл:
      исходное_имя_modified.xlsx

    Добавляет колонки:
      - Цена с Ozon
      - Курс
      - Цена в BYN
    """

    client_id = client_id or OZON_CLIENT_ID
    api_key = api_key or OZON_API_KEY

    if not client_id or not api_key:
        raise ValueError(
            "Не заданы Ozon client_id/api_key. "
            "Передай их аргументами функции или через переменные окружения "
            "OZON_CLIENT_ID и OZON_API_KEY."
        )

    input_path = Path(input_file)
    output_path = build_output_path(input_path)

    if verbose:
        print("=== СТАРТ ===")
        print(f"Входной файл: {input_path}")
        print(f"Выходной файл: {output_path}")
        print(f"Курс BYN к 100 RUB: {byn_rate_per_100_rub}")

    wb = openpyxl.load_workbook(input_path)

    if sheet_name:
        ws = wb[sheet_name]
    elif "Товарный состав" in wb.sheetnames:
        ws = wb["Товарный состав"]
    else:
        ws = wb[wb.sheetnames[0]]

    if verbose:
        print(f"Лист: {ws.title}")
        print(f"Строк всего (с шапкой): {ws.max_row}")
        print(f"Колонок до добавления: {ws.max_column}")

    headers = find_header_map(ws)

    if verbose:
        print("Найденные заголовки:")
        for k, v in headers.items():
            print(f"  {v}: {k}")

    product_id_col = headers.get("ozonid")
    offer_id_col = headers.get("артикул")

    if not product_id_col and not offer_id_col:
        raise ValueError("Не найдена колонка 'OzonID' или 'Артикул' в первой строке листа.")

    if verbose:
        print(f"Колонка OzonID: {product_id_col}")
        print(f"Колонка Артикул: {offer_id_col}")

    price_col, rate_col, byn_col = add_new_columns(
        ws,
        ["Цена с Ozon", "Курс", "Цена в BYN"],
    )

    if verbose:
        print(f"Новые колонки: price={price_col}, rate={rate_col}, byn={byn_col}")

    display_rate = q2(byn_rate_per_100_rub)
    exact_rate = Decimal(str(byn_rate_per_100_rub))
    hundred = Decimal("100")

    # Собираем идентификаторы
    product_ids: list[str] = []
    offer_ids: list[str] = []

    seen_product_ids = set()
    seen_offer_ids = set()

    for row in range(2, ws.max_row + 1):
        if product_id_col:
            pid = parse_product_id(ws.cell(row, product_id_col).value)
            if pid and pid not in seen_product_ids:
                seen_product_ids.add(pid)
                product_ids.append(pid)

        if offer_id_col:
            oid = parse_offer_id(ws.cell(row, offer_id_col).value)
            if oid and oid not in seen_offer_ids:
                seen_offer_ids.add(oid)
                offer_ids.append(oid)

    if verbose:
        print(f"Уникальных OzonID: {len(product_ids)}")
        print(f"Уникальных Артикулов: {len(offer_ids)}")

    # ВАЖНО: теперь делаем ОБА запроса, а не только один
    prices_by_product_id: dict[str, Decimal] = {}
    prices_by_offer_id: dict[str, Decimal] = {}

    if product_ids:
        prices_by_product_id = fetch_prices(
            ids=product_ids,
            id_field="product_id",
            client_id=client_id,
            api_key=api_key,
            verbose=verbose,
        )

    if offer_ids:
        prices_by_offer_id = fetch_prices(
            ids=offer_ids,
            id_field="offer_id",
            client_id=client_id,
            api_key=api_key,
            verbose=verbose,
        )

    found_count = 0
    not_found_count = 0

    for row in range(2, ws.max_row + 1):
        pid = parse_product_id(ws.cell(row, product_id_col).value) if product_id_col else None
        oid = parse_offer_id(ws.cell(row, offer_id_col).value) if offer_id_col else None

        # Курс пишем всегда
        ws.cell(row, rate_col).value = float(display_rate)
        ws.cell(row, rate_col).number_format = "0.00"

        ozon_price: Decimal | None = None
        source = None

        if pid and pid in prices_by_product_id:
            ozon_price = prices_by_product_id[pid]
            source = "product_id"
        elif oid and oid in prices_by_offer_id:
            ozon_price = prices_by_offer_id[oid]
            source = "offer_id"

        if ozon_price is not None:
            byn_price = ((ozon_price * exact_rate) / hundred).quantize(
                Decimal("0.01"),
                rounding=ROUND_HALF_UP,
            )

            ws.cell(row, price_col).value = float(ozon_price)
            ws.cell(row, price_col).number_format = "0.00"

            ws.cell(row, byn_col).value = float(byn_price)
            ws.cell(row, byn_col).number_format = "0.00"

            found_count += 1
        else:
            not_found_count += 1

        if verbose:
            # Показываем первые 20 строк и ВСЕ строки, где цена не найдена
            if row <= 21 or ozon_price is None:
                print(
                    f"Строка {row}: "
                    f"OzonID={pid!r}, Артикул={oid!r}, "
                    f"цена={str(ozon_price) if ozon_price is not None else None}, "
                    f"источник={source}"
                )

    ws.column_dimensions[get_column_letter(price_col)].width = 14
    ws.column_dimensions[get_column_letter(rate_col)].width = 10
    ws.column_dimensions[get_column_letter(byn_col)].width = 14

    wb.save(output_path)

    if verbose:
        print()
        print("=== ГОТОВО ===")
        print(f"Заполнено цен: {found_count}")
        print(f"Не найдено цен: {not_found_count}")
        print(f"Сохранено в: {output_path}")

    return output_path
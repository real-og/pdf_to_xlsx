import os
import re
from typing import Optional, List, Any, Tuple

import pdfplumber
from openpyxl import Workbook
from openpyxl.utils import get_column_letter




HEADERS = [
    "Артикул",
    "Наименование товара",
    "Единица измерения",
    "Колич ество",
    "Цена, руб. коп.",
    "Стоимость, руб. коп",
    "Ставка НДС, %",
    "Сумма НДС, руб. коп.",
    "Стоимость с НДС, руб. коп.",
    "Количество грузовых мест",
    "Масса груза",
    "Примечание",
]

# Достаточно “узнаваемые” слова из шапки товарной таблицы (они могут быть с переносами)
NEEDLE_WORDS = ["Наименование", "Единица", "Колич", "НДС", "Стоимость", "Масса", "Примечание"]

# Часто под шапкой есть строка "1 2 3 4 ..." — её пропускаем
NUM_ROW_RE = re.compile(r"^\s*1\s*$")

# Набор возможных единиц измерения (под твой формат; при необходимости расширишь)
UNITS = {"шт.", "кг.", "л.", "уп.", "компл."}


def norm_cell(s: Optional[Any]) -> str:
    if s is None:
        return ""
    s = str(s).replace("\xa0", " ").replace("\n", " ")
    s = re.sub(r"[ \t]+", " ", s.strip())
    return s


def to_float(s: Any) -> Optional[float]:
    s = norm_cell(s)
    if not s:
        return None
    s = s.replace(" ", "").replace(",", ".")
    try:
        return float(s)
    except ValueError:
        return None


def looks_like_header_row(row: List[Any]) -> bool:
    joined = " ".join(norm_cell(c) for c in row if c).lower()
    hit = sum(1 for w in NEEDLE_WORDS if w.lower() in joined)
    return hit >= 4


def split_article_name(cell0: str) -> Tuple[str, str]:
    """
    В некоторых PDF первая колонка может быть "АРТИКУЛ, Наименование".
    """
    s = norm_cell(cell0)
    if "," in s:
        art, name = s.split(",", 1)
        return art.strip(), " " + name.strip()
    return "", (" " + s.strip()) if s else ""


def table_settings():
    # Совместимо с pdfplumber==0.11.9 (без keep_blank_chars)
    return {
        "vertical_strategy": "lines",
        "horizontal_strategy": "lines",
        "intersection_tolerance": 5,
        "snap_tolerance": 3,
        "join_tolerance": 3,
        "edge_min_length": 10,
        "min_words_vertical": 1,
        "min_words_horizontal": 1,
    }


def normalize_item_row(raw_row: List[Any]) -> Optional[List[str]]:
    """
    Возвращает нормализованную строку товара строго в 12 колонок:
    [Артикул, Наименование, Ед, Кол-во, Цена, Стоимость, Ставка НДС, Сумма НДС, Стоимость с НДС, Места, Масса, Примечание]
    """
    r = [norm_cell(c) for c in raw_row]
    if not any(r):
        return None

    first = norm_cell(r[0]).lower()
    if first.startswith("итого"):
        return None

    # Иногда pdfplumber может вернуть >12 столбцов — склеим хвост в "Примечание"
    if len(r) > 12:
        r = r[:11] + [" ".join(x for x in r[11:] if x)]

    # Вариант 1: PDF уже распознал Артикул и Наименование как 2 отдельные колонки (len==12)
    if len(r) == 12:
        # Ожидаем единицу в колонке 3 (index 2) и числовое количество в колонке 4 (index 3)
        if r[2] in UNITS and to_float(r[3]) is not None:
            r[11] = re.sub(r"\s*/\s*", " / ", r[11]).strip()
            return r
        # иначе — не товарная строка
        return None

    # Вариант 2: PDF дал 11 колонок, где в первой "АРТИКУЛ, Наименование"
    if len(r) == 11:
        # здесь единица в колонке 2 (index 1) и кол-во в колонке 3 (index 2)
        if r[1] in UNITS and to_float(r[2]) is not None:
            art, name = split_article_name(r[0])
            out = [art, name, r[1], r[2], r[3], r[4], r[5], r[6], r[7], r[8], r[9], r[10]]
            out[11] = re.sub(r"\s*/\s*", " / ", out[11]).strip()
            return out

    return None


def extract_items(pdf_path: str) -> List[List[str]]:
    items: List[List[str]] = []
    with pdfplumber.open(pdf_path) as pdf:
        for page in pdf.pages:
            for tbl in (page.extract_tables(table_settings()) or []):
                # ищем шапку таблицы
                header_idx = None
                for i, row in enumerate(tbl[:10]):
                    if looks_like_header_row(row):
                        header_idx = i
                        break
                if header_idx is None:
                    continue

                start = header_idx + 1
                if start < len(tbl) and any(NUM_ROW_RE.match(norm_cell(c)) for c in tbl[start]):
                    start += 1

                for raw_row in tbl[start:]:
                    row = normalize_item_row(raw_row)
                    if row is None:
                        continue
                    # доп. защита: количество должно быть числом
                    if to_float(row[3]) is None:
                        continue
                    items.append(row)

    if not items:
        raise RuntimeError("Не нашёл товарную таблицу в PDF. Возможно другой шаблон или PDF-скан.")

    return items


def write_xlsx(items: List[List[str]], out_path: str) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    # Заголовки
    for c, h in enumerate(HEADERS, start=1):
        ws.cell(row=1, column=c, value=h)

    start_row = 2

    # Товары
    for i, r in enumerate(items):
        row = start_row + i

        art = r[0]
        name = r[1]
        if name and not name.startswith(" "):
            name = " " + name  # как в эталоне
        unit = r[2]

        qty = to_float(r[3])
        price = to_float(r[4])
        cost = to_float(r[5])
        vat_rate = to_float(r[6])
        vat_sum = to_float(r[7])
        cost_vat = to_float(r[8])
        places = to_float(r[9])
        weight = to_float(r[10])
        note = norm_cell(r[11])

        # типы
        if qty is not None and abs(qty - round(qty)) < 1e-9:
            qty = int(round(qty))
        if places is not None and abs(places - round(places)) < 1e-9:
            places = int(round(places))
        if vat_rate is not None and abs(vat_rate - round(vat_rate)) < 1e-9:
            vat_rate = int(round(vat_rate))
        if vat_sum is None:
            vat_sum = 0

        ws.cell(row=row, column=1, value=art)
        ws.cell(row=row, column=2, value=norm_cell(name))
        ws.cell(row=row, column=3, value=unit)
        ws.cell(row=row, column=4, value=qty)
        ws.cell(row=row, column=5, value=price)
        ws.cell(row=row, column=6, value=cost)
        ws.cell(row=row, column=7, value=vat_rate)
        ws.cell(row=row, column=8, value=vat_sum)
        ws.cell(row=row, column=9, value=cost_vat)
        ws.cell(row=row, column=10, value=places)
        ws.cell(row=row, column=11, value=weight)
        ws.cell(row=row, column=12, value=note)

    last_item_row = start_row + len(items) - 1
    total_row = last_item_row + 1
    control_row = last_item_row + 2

    # Считаем “Итого” ЧИСЛАМИ (как в твоём эталонном xlsx)
    sum_qty = sum((to_float(ws.cell(r, 4).value) or 0) for r in range(start_row, last_item_row + 1))
    sum_cost = sum((to_float(ws.cell(r, 6).value) or 0) for r in range(start_row, last_item_row + 1))
    sum_vat = sum((to_float(ws.cell(r, 8).value) or 0) for r in range(start_row, last_item_row + 1))
    sum_cost_vat = sum((to_float(ws.cell(r, 9).value) or 0) for r in range(start_row, last_item_row + 1))
    sum_places = sum((to_float(ws.cell(r, 10).value) or 0) for r in range(start_row, last_item_row + 1))
    sum_weight = sum((to_float(ws.cell(r, 11).value) or 0) for r in range(start_row, last_item_row + 1))

    # Строка "Итого" (предпоследняя)
    ws.cell(row=total_row, column=1, value=" ")
    ws.cell(row=total_row, column=2, value="Итого")
    ws.cell(row=total_row, column=3, value="X")
    ws.cell(row=total_row, column=4, value=int(sum_qty) if abs(sum_qty - round(sum_qty)) < 1e-9 else sum_qty)
    ws.cell(row=total_row, column=5, value="X")
    ws.cell(row=total_row, column=6, value=round(sum_cost, 2))
    ws.cell(row=total_row, column=7, value="X")
    ws.cell(row=total_row, column=8, value=round(sum_vat, 2))
    ws.cell(row=total_row, column=9, value=round(sum_cost_vat, 2))
    ws.cell(row=total_row, column=10, value=int(sum_places) if abs(sum_places - round(sum_places)) < 1e-9 else sum_places)
    ws.cell(row=total_row, column=11, value=round(sum_weight, 3))
    ws.cell(row=total_row, column=12, value="-")

    # "Контроль сумм" (последняя) — формулы как в эталоне
    ws.cell(row=control_row, column=2, value="Контроль сумм")
    ws.cell(row=control_row, column=4, value=f"=SUM(D{start_row}:D{last_item_row})")
    ws.cell(row=control_row, column=6, value=f"=SUM(F{start_row}:F{last_item_row})")
    ws.cell(row=control_row, column=9, value=f"=SUM(I{start_row}:I{last_item_row})")
    ws.cell(row=control_row, column=10, value=f"=SUM(J{start_row}:J{last_item_row})")
    ws.cell(row=control_row, column=11, value=f"=SUM(K{start_row}:K{last_item_row})")

    # Ширины колонок (не критично, но удобно)
    widths = [12, 65, 14, 10, 14, 16, 12, 18, 20, 22, 12, 45]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    wb.save(out_path)


def parse_pdf_to_xlsx(input_pdf, out_name):
    if not os.path.exists(input_pdf):
        raise FileNotFoundError(f"Не найден '{input_pdf}' в текущей папке: {os.getcwd()}")

    items = extract_items(input_pdf)
    write_xlsx(items, out_name)
    print(f"OK: {input_pdf} -> {out_name} (rows: {len(items)})")



